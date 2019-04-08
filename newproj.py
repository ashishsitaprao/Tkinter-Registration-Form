# import open/pyxl and tkinter modules 
from openpyxl import *
import tkinter as tk
from tkinter import *


# globally declare wb and sheet variable 
# opening the existing excel file 
wb = load_workbook('C:\\Users\\ashish\\Desktop\\excel.xlsx') 

# create the sheet object 
sheet = wb.active 



def excel(): 
	
	# resize the width of columns in 
	# excel spreadsheet 
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 20
	sheet.column_dimensions['C'].width = 20
	sheet.column_dimensions['D'].width = 50
	sheet.column_dimensions['E'].width = 50
	sheet.column_dimensions['F'].width = 20
	sheet.column_dimensions['G'].width = 20
	sheet.column_dimensions['H'].width = 20
	sheet.column_dimensions['I'].width = 20
	sheet.column_dimensions['J'].width = 20
	sheet.column_dimensions['K'].width = 20
	sheet.column_dimensions['L'].width = 20
	sheet.column_dimensions['M'].width = 20
	sheet.column_dimensions['N'].width = 20
	sheet.column_dimensions['O'].width = 20

	# write given data to an excel spreadsheet 
	# at particular location 
	sheet.cell(row=1, column=1).value = "Name"
	sheet.cell(row=1, column=2).value = "Branch"
	sheet.cell(row=1, column=3).value = "Year"
	sheet.cell(row=1, column=4).value = "Email ID"
	sheet.cell(row=1, column=5).value = "Contact Number"
	sheet.cell(row=1, column=6).value = "Zodies"
	sheet.cell(row=1, column=7).value = "Laser Tag"
	sheet.cell(row=1, column=8).value = "Debate"
	sheet.cell(row=1, column=9).value = "Kidalt"
	sheet.cell(row=1, column=10).value = "Face Painting"
	sheet.cell(row=1, column=11).value = "Fine Art"
	sheet.cell(row=1, column=12).value = "Vocal Wizard"
	sheet.cell(row=1, column=13).value = "Mr/Mrs Zodiac"
	sheet.cell(row=1, column=14).value = "Glow Cricket"
	sheet.cell(row=1, column=15).value = "Open Mic"

# clear the content of text entry box
def clear(): 
	
	name_field.delete(0, END)  
	email_field.delete(0, END) 
	mobile_field.delete(0, END)
	c1.deselect()
	c2.deselect()
	c3.deselect()
	c4.deselect()
	c5.deselect()
	c6.deselect()
	c7.deselect()
	c8.deselect()
	c9.deselect()
	c10.deselect()
	

# Function to take data from GUI 
# window and write to an excel file 
def insert(): 
	
	# if user not fill any entry 
	# then print "empty input" 
	if (name_field.get() == "" and
		email_field.get() == "" and
		mobile_field.get() == ""): 
			
		print("empty input") 

	else: 

		# assigning the max row and max column 
		# value upto which data is written 
		# in an excel sheet to the variable 
		current_row = sheet.max_row 
		current_column = sheet.max_column 

		# get method returns current text 
		# as string which we write into 
		# excel spreadsheet at particular location 
		sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
		sheet.cell(row=current_row + 1, column=2).value = tkvar_branch.get() 
		sheet.cell(row=current_row + 1, column=3).value = tkvar_year.get() 
		sheet.cell(row=current_row + 1, column=4).value = email_field.get() 
		sheet.cell(row=current_row + 1, column=5).value = mobile_field.get() 
		sheet.cell(row=current_row + 1, column=6).value = var1.get() 
		sheet.cell(row=current_row + 1, column=7).value = var2.get() 
		sheet.cell(row=current_row + 1, column=8).value = var3.get() 
		sheet.cell(row=current_row + 1, column=9).value = var4.get() 
		sheet.cell(row=current_row + 1, column=10).value = var5.get() 
		sheet.cell(row=current_row + 1, column=11).value = var6.get() 
		sheet.cell(row=current_row + 1, column=12).value = var7.get() 
		sheet.cell(row=current_row + 1, column=13).value = var8.get() 
		sheet.cell(row=current_row + 1, column=14).value = var9.get() 
		sheet.cell(row=current_row + 1, column=15).value = var10.get() 

		# save the file 
		wb.save('C:\\Users\\ashish\\Desktop\\excel.xlsx') 

		# set focus on the name_field box 
		name_field.focus_set() 

		# call the clear() function 
		clear() 


#MAIN CODE

root = tk.Tk()


root.configure(background = 'gray25')

root.title("Zodiac Registration Form")

root.geometry("700x500")



heading=tk.Label(root,text="Registration Form",bg='gray25',font ='Times 14',fg='white')


name = tk.Label(root,text="Name",bg='gray25',fg='white')
branch=tk.Label(root,text="Branch",bg='gray25',fg='white')
year=tk.Label(root,text="Year",bg='gray25',fg='white')
email=tk.Label(root,text="Email ID",bg='gray25',fg='white')
mobile=tk.Label(root,text="Mob. No.",bg='gray25',fg='white')

heading.place(x=285,y=15)
name.place(x=40,y=60)
branch.place(x=40,y=140)
year.place(x=40,y=180)
email.place(x=40,y=100)
mobile.place(x=40,y=220)

name_field = tk.Entry(root) 
name_field.config(bg='gray28',fg='white')
mobile_field = tk.Entry(root) 
mobile_field.config(bg='gray28',fg='white')
email_field = tk.Entry(root)  
email_field.config(bg='gray28',fg='white')

name_field.place(x=100, y=60)
email_field.place(x=100,y=100)
mobile_field.place(x=100,y=220)



# Branch And Year Choices
tkvar_branch=tk.StringVar(root)
tkvar_year=tk.StringVar(root)
choices_branch={'COMP','EXTC','INSTRU','MECH','IT'}
choices_year={'First Year','Second Year','Third Year','Fourth Year'}
tkvar_branch.set('COMP')
tkvar_year.set('Second Year')
drop_branch=tk.OptionMenu(root,tkvar_branch,*choices_branch)
drop_year=tk.OptionMenu(root,tkvar_year,*choices_year)

#Setting Color to OPtionMenu
drop_branch.config(bg='gray26',fg='white')
drop_year.config(bg='gray26',fg='white')

drop_branch.place(x=100,y=140)
drop_year.place(x=100,y=180)


#Checkboxes For Different Events
var1=tk.StringVar()
c1 = tk.Checkbutton(root, text="Zodies", variable=var1,onvalue="YES", offvalue="NO",bg='gray25',fg='black')
var2=tk.StringVar()
c2 = tk.Checkbutton(root, text="Laser Tag", variable=var2,onvalue="YES", offvalue="NO",bg='gray25',fg='black')
var3=tk.StringVar()
c3 = tk.Checkbutton(root, text="Debate", variable=var3,onvalue="YES", offvalue="NO",bg='gray25',fg='black')
var4=tk.StringVar()
c4 = tk.Checkbutton(root, text="Kidalt", variable=var4,onvalue="YES", offvalue="NO",bg='gray25',fg='black')
var5=tk.StringVar()
c5 = tk.Checkbutton(root, text="Face Painting", variable=var5,onvalue="YES", offvalue="NO",bg='gray25',fg='black')
var6=tk.StringVar()
c6 = tk.Checkbutton(root, text="Fine Art", variable=var6,onvalue="YES", offvalue="NO",bg='gray25',fg='black')
var7=tk.StringVar()
c7 = tk.Checkbutton(root, text="Vocal Wizard", variable=var7,onvalue="YES", offvalue="NO",bg='gray25',fg='black')
var8=tk.StringVar()
c8 = tk.Checkbutton(root, text="Mr/Mrs Zodiac", variable=var8,onvalue="YES", offvalue="NO",bg='gray25',fg='black')
var9=tk.StringVar()
c9 = tk.Checkbutton(root, text="Glow Cricket", variable=var9,onvalue="YES", offvalue="NO",bg='gray25',fg='black')
var10=tk.StringVar()
c10 = tk.Checkbutton(root, text="Open Mic", variable=var10,onvalue="YES", offvalue="NO",bg='gray25',fg='black')

c1.deselect()
c2.deselect()
c3.deselect()
c4.deselect()
c5.deselect()
c6.deselect()
c7.deselect()
c8.deselect()
c9.deselect()
c10.deselect()

c1.place(x=40 ,y=265)
c2.place(x=160 ,y=265)
c3.place(x=280 ,y=265)
c4.place(x=400 ,y=265)
c5.place(x=520,y=265)
c6.place(x=40,y=310)
c7.place(x=160,y=310)
c8.place(x=280,y=310)
c9.place(x=400,y=310)
c10.place(x=520,y=310)

def new_page():
	popup = Toplevel(bg='gray25')
	popup.geometry('300x200')
	popup.title("Succesful Registration")
	info_display = Label(popup,text="Thank You For Registering",bg='gray25',fg='white')
	info_display.place(x=90,y=70)
	dismiss_info=Button(popup,text="Close",command=popup.destroy,bg="#3700B3",fg="white")
	dismiss_info.place(x=130,y=150)

excel()

def callf() :
	insert()
	new_page()



#Register Button (Saving Data to Excel Sheet)
submit = tk.Button(root, text="Submit", fg="white",bg="#3700B3",command=callf)
submit.place(x=300,y=450)
root.mainloop()